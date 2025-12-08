from __future__ import annotations

from typing import List, Dict, Any, Tuple
import io
import csv
from openpyxl import load_workbook

try:
    import xlrd  # for .xls legacy support
except Exception:  # pragma: no cover
    xlrd = None


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _read_csv(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    text = file_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows_iter = iter(reader)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], []
    header_order = [str(h) for h in header]
    rows: List[Dict[str, Any]] = []
    for r in rows_iter:
        record = {}
        for idx, col in enumerate(header_order):
            record[col] = _stringify(r[idx]) if idx < len(r) else ""
        rows.append(record)
    return header_order, rows


def _read_xlsx(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=True)
    ws = wb.active
    first = True
    header_order: List[str] = []
    rows: List[Dict[str, Any]] = []
    for row_cells in ws.iter_rows(values_only=False):
        values = [_stringify(c.value) for c in row_cells]
        if first:
            header_order = [str(h) for h in values]
            first = False
            continue
        if not header_order:
            continue
        record = {}
        for idx, col in enumerate(header_order):
            record[col] = values[idx] if idx < len(values) else ""
        rows.append(record)
    return header_order, rows


def _read_xls(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    if xlrd is None:
        raise ValueError(".xls support requires xlrd; please install xlrd==1.2.0")
    book = xlrd.open_workbook(file_contents=file_bytes)
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        return [], []
    header_order = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
    rows: List[Dict[str, Any]] = []
    for r in range(1, sheet.nrows):
        record = {}
        for c in range(sheet.ncols):
            record[header_order[c]] = _stringify(sheet.cell_value(r, c))
        rows.append(record)
    return header_order, rows


def read_file_preserve_text(filename: str, file_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        return _read_csv(file_bytes)
    if lower_name.endswith(".xlsx"):
        return _read_xlsx(file_bytes)
    if lower_name.endswith(".xls"):
        return _read_xls(file_bytes)
    raise ValueError("Unsupported file type. Please upload CSV or Excel files.")


